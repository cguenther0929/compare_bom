"""
FILE: compare_bom_xlsx.py

PURPOSE: 
To compare two BOMs (i.e. between IFS and Engineering).  

The comparison algorithm will first look at ever QPN 
in the engineering BOM and compare against IFS.  Secondly 
it will look at every QPN in the IFS BOM and compare against
the ENG BOM.  This helps assure there aren't parts listed on
one BOM that aren't on the other.  

AUTHOR: 
Clinton G. 

TODO: Nothing

"""
import sys
import random
import time
import csv
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

# ----------------------------------------------------------------------- #
# Regular Expression Strings
# ----------------------------------------------------------------------- #
qpn_re 		= "(QPN)|(COMPONENT.?PART)"
mfgpn_re 	= "(MFG.?PN)"										# To match MFGPN or MFG PN (will ignore case)
mfg_re 		= "(MFG)|(MANUFACTURER)"			
des_re 		= "(DES)|(DESCRIPTION)|(Part.?Description)"
ref_re		= "(REF)|(REF.DES)|(REFERENCE)|(NOTES)"				# IFS BOMs often put this information in the NOTES column
qty_re		= "(QTY)|(QUANTITY)|(Qty.{1,20})"
uom_re		= "(UOM)|(UNIT OF MEASURE)"
cr1_re		= "(CR1)"
cr1pn_re	= "(CR1PN)"
notes_re	= "(NOTES)"



## DEFINE VRIABLES ##
#####################
MFGPN_col 	= 0								# Column number containing the MFGPN
QPN_col 	= 0								# Column number containing QPN
MFG_col 	= 0								# Column location for manufacturer part number
DES_col 	= 0 							# Column location for description part number
QTY_col 	= 0 							# Column location for quantity field
UOM_col 	= 0 							# Column location for UOM field
CR1_col		= 0								# Column location for supplier name
CR1PN_col	= 0								# Column location for supplier's PN
NOTE_col 	= 0 							# Column location for "notes" field
BOM_HEADER 	= ["QPN","QTY","DES","REF"]		# The IFS BOM dictates this

# -------------------------------------- #
# Dictionaries
# -------------------------------------- #
dict_eng_bom	= {}
dict_ifs_bom	= {}


# -------------------------------------- #
# Boolean Flags
# -------------------------------------- #
bom_is_ifs				= False		# True if IFS BOM / False if ENG BOM
flag_header_detecetd 	= False		# Set to true as soon as we detect header data in one of the rows
sheet_valid		= False		# Flag that tells application if a sheet contains valid data or not


data_start 				= 0			# This is the row where the data starts
search_header 	= []		# Set equal to BOM_HEADER and pop elements until we find all the colums we're looking for
header 			= []		# This array will define the column locations for the header
qpn 			= []        # Pull in all QPNs into a list. This will make them easier to work with later
asso 			= []       	# Pull in all associations into a list. This will make them easier to work with later
qty 			= []        # Pull in all QTYs into a list. This will make them easier to work with later
uom 			= []        # Pull in all UOM values into a list. This will make them easier to work with later
des 			= []		# Pull in all Descriptions into a list. This will make them easier to work with later
ref 			= []		# Pull all reference values into a list. This will make them easier to work with later
mfg 			= []		# Pull in all Manufactures into a list. This will make them easier to work with later
mfgpn 			= []		# Pull in all Manufacturing Part Numbers into a list. This will make them easier to work with later
cr1 			= []		# Pull in all suppler names into a list. This will make them easier to work with later
cr1pn 			= []		# Pull in all supplier pn's into a list. This will make them easier to work with later
notes 			= []		# Pull all note values into a list. This will make them easier to work with later


# -------------------------------------- #
# Local Methods
# -------------------------------------- #
def debugbreak():
	while(1):
		pass
		
def clean_value(textin):
	temptext = textin
	logging.info("Text entered into method clean value: " + str(temptext))
	temptext = temptext.lstrip('text:u\'')     	# Remove the initial part of the string that we don't need 'text:u'   
	temptext = temptext.lstrip("b\'")     	# Remove the initial part of the string that we don't need 'text:u'   
	temptext = temptext.replace("'","")			# Remove single quote marks from value
	temptext = temptext.strip()					# Remove only leading and trailing white spaces
	if(temptext.find("number:") != -1):
		temptext = temptext.replace("number:","")			#This will remove any and all white spaces
	if(temptext.find("mpty:") != -1):
		temptext = temptext.replace("mpty:","")			#This will remove any and all white spaces
	return temptext
	
def clean_des(textin):
	temptext = textin
	temptext = temptext.lstrip('text:u\'')     #Remove the initial part of the string that we don't need 'text:u'  
	temptext = temptext.replace("'","")			#This will remove any and all white spaces
	if(temptext.find("mpty:") != -1):
		temptext = temptext.replace("mpty:","")			#This will remove any and all white spaces
	return temptext
def pause():
	user_input=input("Press any key to exit...")
	sys.exit(0)

# -------------------------------------- #
# Setup Logging
# -------------------------------------- #
logging.basicConfig(
    filename = 'compare_bom.log',
    level = logging.DEBUG,
    format =' %(asctime)s -  %(levelname)s - %(message)s',
	filemode = 'w'
)
	   
#****************************************************************************** 
#******************************  ---MAIN---  **********************************
#******************************************************************************   
if __name__ == '__main__':

	path = os.getcwd()
	# Find path/dirs/files
	for (path, dirs, files) in os.walk(path):
		path
		dirs
		files

	print ("Files found in directory: ", str(len(files)))
	print ("File names: ", files)

	# ----------------------------------------------------------------------- #
	# Iterate through files
	# ----------------------------------------------------------------------- #
	for i in range(len(files)):
		
		# ----------------------------------------------------------------------- #
		# Only open files having the proper extension 
		# ----------------------------------------------------------------------- #
		if(files[i].upper().endswith(".XLSX")):
			
			print ("Opening file: ", files[i])
			wb = load_workbook(filename = files[i])     # Open the workbook that we are going to parse though 
			ws = wb.sheetnames             				# Grab the names of the worksheets -- I believe this line is critical.
			
			# Each BOM / workbook shall only contain one sheet with 
			# BOM data.  However, often times BOMs include a revision sheet / etc.,
			# thus this script shall be intelligent enough to properly omit 
			# revision/changelog/etc. sheets.  
			num_sheets = len(ws)						# This is the number of sheets

			# ----------------------------------------------------------------------- #
			# Determine BOM Origin (ENG or IFS)
			# ----------------------------------------------------------------------- #
			# TODO need to search file string to determine if ENG or IFS
			if ((files[i].find("ENG") != -1) or (files[i].find("eng") != -1)):
				bom_is_ifs = False
			elif ((files[i].find("IFS") != -1) or (files[i].find("ifs") != -1)):
				bom_is_ifs = True
			else:
				logging.info("Asking user to distinguish ENG vs. IFS BOM.")
				while(True):
					temp_input = input("Is this the ENG BOM (y/n): ")   
					if(temp_input == 'y'):
						bom_is_ifs = False
						break
					elif(temp_input == 'n'):
						bom_is_ifs = True
						break
					else:
						pass
			
			
			print ("\n\n===============================================")
			print ("===============================================")
			print ("File opened: ", str(files[i]))
			
			if(bom_is_ifs):
				print("This has been detected as the IFS BOM.")
				logging.info("This has been detected as the IFS BOM.")
			else:
				print("This has been detected as the ENG BOM.")
				logging.info("This has been detected as the ENG BOM.")

			print ("The number of worksheets is: ", str(num_sheets))
			print ("Worksheet names: ", ws)
			print ("===============================================")

			# ----------------------------------------------------------------------- #
			# Iterate through all sheets
			# ----------------------------------------------------------------------- #
			for sh in range (num_sheets):
				sheet_valid = False
				
				current_sheet = wb[ws[sh]]
				
				print("Now operating on worksheet: ", ws[sh])
				
				num_rows = current_sheet.max_row     		
				num_cols = current_sheet.max_column 		

				# ----------------------------------------------------------------------- #
				# Iterate through every row on current sheet
				# ----------------------------------------------------------------------- #
				for r in range (1,num_rows + 1):					# Find the header locations. Excel starts counting at one
					search_header = BOM_HEADER.copy()						# Load up headers we need to search for
					print ("Search header before starting: ", search_header)
					
					flag_header_detecetd = False
					# ----------------------------------------------------------------------- #
					# Iterate over columns of selected row
					# ----------------------------------------------------------------------- #
					for c in range (1,num_cols + 1):				# Excel starts counting at 1
						
						temptext = str(str(current_sheet.cell(row = r, column=c).value).encode(encoding = 'UTF-8',errors = 'strict'))                
						temptext = temptext.lstrip("text:u\'")     	# Clean up the extra garbage on text 
						temptext = temptext.lstrip("b\'")     		
						temptext = temptext.rstrip("\'")     		
						temptext = temptext.replace(" ","")			# Remove any and all white spaces 
						logging.info("Text extracted from cell: " + temptext)
						# print ("****DEBUG Text Extracted: ", temptext)
						# print ("****DEBUG Current column number: ", str(c))


						
						if(re.fullmatch(qpn_re,temptext,re.IGNORECASE)):
							QPN_col = c
							search_header.remove("QPN")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found QPN")
						
						elif(re.fullmatch(des_re,temptext,re.IGNORECASE)):		#Look for Description
							DES_col = c
							search_header.remove("DES")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found DES")
						
						elif(re.fullmatch(ref_re,temptext,re.IGNORECASE)):		#Look for Description
							REF_col = c
							search_header.remove("REF")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found REF")
						
						elif(re.fullmatch(qty_re,temptext,re.IGNORECASE)):		#Look for Quantity field.  
							QTY_col = c
							search_header.remove("QTY")
							logging.info("Found header: " + temptext)
							logging.info("Still Looking For: " + str(search_header))
							# print("**** DEBUG found QTY")

					if( (len(search_header) == 0) ):		# Found all header fields
						sheet_valid = True
						data_start = r + 1			# Plenty of confidence at this point that we've found data start
						print ("Data appears to start on row: ", data_start)
						logging.info("Data appears to start on row: " + str(data_start))
						
						print( 	"Sample data in start row: ", clean_value(str(str(current_sheet.cell(row = data_start, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict'))),' ', 
								clean_value(str(str(current_sheet.cell(row = data_start, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
								clean_value(str(str(current_sheet.cell(row = data_start, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							)
						break

					elif((r == 10) and (len(search_header) > 0) and sh < num_sheets):
						sheet_valid = False
						print ("* File: ", str(files[i]), "Invalid Sheet: ", str(ws[sh]), " -- did not find headers: ", search_header)
						break
					
				if(sheet_valid):
					print ("QPN column found to be: ", 			str(QPN_col))		
					print ("QTY column found to be: ", 			str(QTY_col))
					print ("Description column found to be: ", 	str(DES_col))		
					print ("Reference column found to be: ", 	str(REF_col))		
					
					header = [QPN_col,DES_col,REF_col,QTY_col]
					header_values = ["QPN","DES","REF","QTY","NOTES"]
					
					# Now iterate through all rows of the current sheet and populate the data lists
					blank_row_count = 0		# Reset number of blank rows detected.  When three in a row are detected, break out of the loop. 
					for r in range (data_start,num_rows + 1):
						
						
						# If multiple columns are blank, break out of this loop for these are empty cells
						if( ( len(clean_des(str(str(current_sheet.cell(row = r, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) and
							( len(clean_des(str(str(current_sheet.cell(row = r, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) and
						 	( len(clean_des(str(str(current_sheet.cell(row = r, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) and
							( len(clean_des(str(str(current_sheet.cell(row = r, column=QTY_col).value).encode(encoding = 'UTF-8',errors = 'strict')))) <= 1) ):
							
							blank_row_count += 1				# Increase value of blank row count
							print ("Blank row detected at row (", r, ")")
						
						else:
							
							blank_row_count = 0					
							print( 	'Sample data, current row: ', 
									clean_value(str(str(current_sheet.cell(row = r, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = r, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = r, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict'))), ' ', 
									clean_value(str(str(current_sheet.cell(row = r, column=QTY_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
								)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=QPN_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							qpn.append(current_value)			
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=DES_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							des.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=REF_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							ref.append(current_value)
							
							current_value = clean_value(str(str(current_sheet.cell(row = r, column=QTY_col).value).encode(encoding = 'UTF-8',errors = 'strict')))
							if current_value == "None":
								current_value = ""
							qty.append(current_value)
							
						if(blank_row_count >= 3):
							break								# Too many blank rows detected, so break out of the loop.  
	
		# ----------------------------------------------------------------------- #
		# If sheet is valid, and before moving to next file
		# build the dictionary for the respective BOM
		# ----------------------------------------------------------------------- #

		if(sheet_valid):
			# Construct dictionary 
			if(bom_is_ifs):
				for i in range (0,len(qpn)):				
					dict_ifs_bom[qpn[i]] = (des[i],ref[i],qty[i])
			else:
				for i in range (0,len(qpn)):				
					dict_eng_bom[qpn[i]] = (des[i],ref[i],qty[i])

		# ----------------------------------------------------------------------- #
		# Lists shall be cleared before moving onto the 
		# next file, as a different dictionary will need to populated
		# ----------------------------------------------------------------------- #
		qpn.clear()
		des.clear()
		ref.clear()
		qty.clear()
		

					
	# ----------------------------------------------------------------------- #
	# Main Loop
	# Dictionaries have been built, and it is now time to compare
	# between the two BOMs
	# ----------------------------------------------------------------------- #

	
	# ----------------------------------------------------------------------- #
	# Iterate through every QPN in the ENG BOM
	# and compare against IFS
	# ----------------------------------------------------------------------- #
	print("\n\n================================================")
	print("================================================")
	print("All QPN Matches")
	logging.info("================================================")
	logging.info("================================================")
	logging.info("All Matches")
	
	for key in dict_eng_bom:
		if(key in dict_ifs_bom):
			print("QPN: ", key, " -- in ENG and IFS BOM")
			
			print("\tENG/IFS DES:\t", dict_eng_bom[key][0]," | ",dict_ifs_bom[key][0])
			print("\tENG/IFS QTY:\t", dict_eng_bom[key][2]," | ",dict_ifs_bom[key][2])
			print("\tENG/IFS REF:\t", dict_eng_bom[key][1]," | ",dict_ifs_bom[key][1])
		
	print("\n================================================")
	print("================================================")
	print("In Engineering but not in IFS")
	logging.info("================================================")
	logging.info("================================================")
	logging.info("In Engineering but not in IFS")
		
	for key in dict_eng_bom:
		if (key not in dict_ifs_bom):
			print("QPN ", key, " -- is in ENG but NOT in IFS.")


	print("\n================================================")
	print("================================================")
	print("In IFS, but not in Engineering")
	logging.info("================================================")
	logging.info("================================================")
	logging.info("In IFS, but not in Engineering")

	for key in dict_ifs_bom:
		if(key not in dict_eng_bom):
			print("QPN: ", key, " -- is in IFS but NOT in ENG.")
		# else:
		# 	print("Key -- ", key, " -- is NOT in the ENG BOM.")
			
	print("\n")
	
	
	
	# ----------------------------------------------------------------------- #
	# Create the comparison BOM
	# ----------------------------------------------------------------------- #
	print("\n================================================")
	print("================================================")
	print("Creating comparison BOM")
	logging.info("================================================")
	logging.info("================================================")
	logging.info("Creating comparison BOM")
	
	NewBook = Workbook()
	NewSheet = NewBook.active
	NewSheet.title = "Comparison Data"

	current_sheet = wb[ws[sh]]

	# ----------------------------------------------------------------------- #
	# Format column widths
	# ----------------------------------------------------------------------- #
	NewSheet.column_dimensions['A'].width = 25			# QPN
	NewSheet.column_dimensions['B'].width = 25
	NewSheet.column_dimensions['C'].width = 5			# Dash
	NewSheet.column_dimensions['D'].width = 50			# Description
	NewSheet.column_dimensions['E'].width = 50			
	NewSheet.column_dimensions['F'].width = 5			# Dash
	NewSheet.column_dimensions['G'].width = 30			# REF
	NewSheet.column_dimensions['H'].width = 30			
	NewSheet.column_dimensions['I'].width = 5			# Dash
	NewSheet.column_dimensions['J'].width = 15			# QTY
	NewSheet.column_dimensions['K'].width = 15
	
	
	comparison_bom_header = ["IFS QPN","ENG QPN","-","IFS DES","ENG DES","-","IFS REF","ENG REF","-","IFS QTY","ENG QTY"]
	comparison_bom_col_offsets = {"IFS_QPN":1,"ENG_QPN":2,"IFS_DES":4,"ENG_DES":5,"IFS_REF":7,"ENG_REF":8,"IFS_QTY":10,"ENG_QTY":11}
	current_row_counter = 1
	# ----------------------------------------------------------------------- #
	# Write the header values -- first column
	# starts at 0, not 1
	# ----------------------------------------------------------------------- #
	for i in range (1,len(comparison_bom_header)+1):
		NewSheet.cell(row=current_row_counter,column=i).value = comparison_bom_header[i-1]
	current_row_counter = current_row_counter + 1 
	
	# ----------------------------------------------------------------------- #
	# Iterate through every QPN in the ENG BOM
	# and compare against IFS
	# ----------------------------------------------------------------------- #
	logging.info("================================================")
	logging.info("Writing to BOM the components that match")
	
	NewSheet.cell(row=current_row_counter,column=1).value = "These QPNs match between IFS and ENG"
	current_row_counter = current_row_counter + 1 
	
	for key in dict_eng_bom:
		if(key in dict_ifs_bom):
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_QPN"]).value = key
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_QPN"]).value = key
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_DES"]).value = dict_ifs_bom[key][0]
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_DES"]).value = dict_eng_bom[key][0]
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_REF"]).value = dict_ifs_bom[key][1]
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_REF"]).value = dict_eng_bom[key][1]
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_QTY"]).value = dict_ifs_bom[key][2]
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_QTY"]).value = dict_eng_bom[key][2]
			
			current_row_counter = current_row_counter + 1 
			
	current_row_counter = current_row_counter + 2 
			
			
	# ----------------------------------------------------------------------- #
	# Iterate through every QPN in the ENG BOM
	# and compare against IFS
	# ----------------------------------------------------------------------- #
	logging.info("================================================")
	logging.info("Writing to BOM the components in ENG BOM but not in IFS")
	
	NewSheet.cell(row=current_row_counter,column=1).value = "These QPNs are in ENG, but NOT in IFS"
	current_row_counter = current_row_counter + 1 
	
	for key in dict_eng_bom:
		if (key not in dict_ifs_bom):
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_QPN"]).value = key
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_DES"]).value = dict_eng_bom[key][0]
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_REF"]).value = dict_eng_bom[key][1]
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["ENG_QTY"]).value = dict_eng_bom[key][2]
	
			current_row_counter = current_row_counter + 1 
	
	current_row_counter = current_row_counter + 2 


	# ----------------------------------------------------------------------- #
	# Iterate through every QPN in the IFS BOM
	# and compare against ENG
	# ----------------------------------------------------------------------- #
	logging.info("================================================")
	logging.info("Writing to BOM the components in IFS BOM but not in ENG")
	
	NewSheet.cell(row=current_row_counter,column=1).value = "These QPNs are in IFS, but NOT in ENG"
	current_row_counter = current_row_counter + 1 
	
	for key in dict_ifs_bom:
		if(key not in dict_eng_bom):
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_QPN"]).value = key
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_DES"]).value = dict_ifs_bom[key][0]
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_REF"]).value = dict_ifs_bom[key][1]
			
			NewSheet.cell(row=current_row_counter,column=comparison_bom_col_offsets["IFS_QTY"]).value = dict_ifs_bom[key][2]
	
			current_row_counter = current_row_counter + 1 
	
	# ----------------------------------------------------------------------- #
	# Close comparison workbook
	# ----------------------------------------------------------------------- #
	NewBook.save(filename = "Comparison_Results.xlsx")
	print ("\n")
	null=input("Press any key to close...")